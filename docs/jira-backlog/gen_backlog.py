"""Generate a Jira backlog .xlsx in the exact style of the IAM System19 reference.

Usage: build a `rows` list of dicts and a `meta` dict, then call build().
Each row dict: issue_type, summary, epic_name, epic_link, priority,
story_points (int or None), sprint, labels, status, description.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---- palette captured from the reference workbook ----
NAVY = "1F3564"
EPIC_FILL = "D6E4F7"
STORY_FILL = "F5F8FF"
STATUS_STYLE = {
    "Done":    ("E2EFDA", "375623"),
    "Partial": ("FFF2CC", "7F6000"),
    "To Do":   ("FCE4D6", "843C0C"),
}
HEADERS = ["Issue Type", "Summary", "Epic Name", "Epic Link", "Priority",
           "Story Points", "Sprint", "Labels", "Status", "Description"]
WIDTHS = {"A": 12, "B": 50, "C": 28, "D": 12, "E": 10,
          "F": 10, "G": 10, "H": 26, "I": 10, "J": 62}

thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex6):
    return PatternFill("solid", fgColor=hex6)


def build(rows, out_path, backlog_sheet_name="EVS Backlog"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = backlog_sheet_name

    # header
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    r = 2
    for row in rows:
        is_epic = row["issue_type"] == "Epic"
        rowfill = EPIC_FILL if is_epic else STORY_FILL
        vals = [row["issue_type"], row["summary"], row["epic_name"], row["epic_link"],
                row["priority"], row.get("story_points"), row["sprint"], row["labels"],
                row["status"], row["description"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = _fill(rowfill)
            cell.border = BORDER
            cell.font = Font(bold=is_epic)
            wrap = c in (2, 3, 8, 10)
            cell.alignment = Alignment(wrap_text=wrap, vertical="center",
                                       horizontal="center" if c in (1, 5, 6, 7, 9) else "left")
        # status colour override (cols I=9)
        sfill, sfont = STATUS_STYLE.get(row["status"], (rowfill, "000000"))
        scell = ws.cell(row=r, column=9)
        scell.fill = _fill(sfill)
        scell.font = Font(bold=True, color=sfont)
        ws.row_dimensions[r].height = 42 if is_epic else 36
        r += 1

    # ---- Sprint Summary sheet ----
    ss = wb.create_sheet("Sprint Summary")
    ss_headers = ["Sprint", "Stories", "Story Points", "Done", "% Complete"]
    for c, h in enumerate(ss_headers, start=1):
        cell = ss.cell(row=1, column=c, value=h)
        cell.fill = _fill(NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for col in "ABCDE":
        ss.column_dimensions[col].width = 14

    # aggregate stories (exclude epics) by sprint
    from collections import defaultdict
    agg = defaultdict(lambda: {"stories": 0, "points": 0, "done": 0})
    for row in rows:
        if row["issue_type"] == "Epic":
            continue
        s = row["sprint"]
        agg[s]["stories"] += 1
        agg[s]["points"] += row.get("story_points") or 0
        if row["status"] == "Done":
            agg[s]["done"] += 1

    def sprint_key(name):
        try:
            return int(name.split()[-1])
        except Exception:
            return 999

    rr = 2
    for sprint in sorted(agg, key=sprint_key):
        a = agg[sprint]
        pct = f"{round(100*a['done']/a['stories'])}%" if a["stories"] else "0%"
        for c, v in enumerate([sprint, a["stories"], a["points"], a["done"], pct], start=1):
            cell = ss.cell(row=rr, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c != 1 else "left")
        rr += 1

    wb.save(out_path)
    # quick stats
    n_epic = sum(1 for x in rows if x["issue_type"] == "Epic")
    n_story = sum(1 for x in rows if x["issue_type"] == "Story")
    print(f"Wrote {out_path}: {n_epic} epics, {n_story} stories")


if __name__ == "__main__":
    # smoke test
    demo = [{"issue_type": "Epic", "summary": "Test", "epic_name": "T", "epic_link": "T",
             "priority": "High", "story_points": None, "sprint": "Sprint 1",
             "labels": "x", "status": "Partial", "description": "d"}]
    build(demo, "/tmp/_smoke.xlsx")
